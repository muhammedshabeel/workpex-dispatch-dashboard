from __future__ import annotations

import csv
import io
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADERS = [
    "client_order_ref", "customer_name", "partner_id", "whatsapp_no", "source_id",
    "Pricelist Name", "street_no", "building_no", "zone_id (governarate)",
    "wilayat_id", "city_id", "order_line/product_id", "order_line/product_uom",
    "order_line/price_unit", "order_line/product_uom_qty", "remarks", "Agent Name", "Source",
    "Payment Method",
]

SETTINGS = {
    "qatar": {
        "countries": {"qatar", "qa"},
        "code": "974",
        "prefix": "EMQ",
        "source_id": "SCENT PASSION - QATAR",
        "pricelist": "Public Pricelist",
        "location_file": "qatar_locations.csv",
        "zone_field": "ZONE",
    },
    "bahrain": {
        "countries": {"bahrain", "bh"},
        "code": "973",
        "prefix": "EMB",
        "source_id": "SCENT PASSION - BAHRAIN",
        "pricelist": "Default BHD pricelist",
        "location_file": "bahrain_locations.csv",
        "zone_field": "BLOCK",
    },
}

ALIASES = {
    "name": ["Lead Name", "Customer Name", "Name", "First Name"],
    "phone1": ["Primary Phone", "Phone 1", "Phone1", "Phone", "Mobile", "Mobile No"],
    "phone2": ["Secondary Phone", "Phone 2", "Phone2", "Alternate Phone", "WhatsApp Number"],
    "country": ["Country"],
    "street": ["Street", "Address", "Address 1"],
    "building": ["Building No", "Building Number", "Building", "Flat/Villa No"],
    "product": ["Product", "Product Name"],
    "qty": ["QTY", "Quantity"],
    "product2": ["PRODUCT 2", "Product 2"],
    "qty2": ["QTY OF PRODUCT 2", "Quantity of Product 2", "QTY 2", "Qty 2"],
    "amount": ["Actual Amount", "Forecasted Amount", "Amount", "COD Amount"],
    "payment": ["Payment Method", "Payment"],
    "remarks": ["Lead Description", "Remarks", "Notes"],
    "reference": ["National Code", "Reference No", "Order ID"],
    "agent_name": ["Assigned", "Assigned User"],
    "source": ["Source"],
}


@dataclass(frozen=True)
class GCCResult:
    dataframe: pd.DataFrame
    workbook_bytes: bytes
    exported_rows: int


def _clean(value) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def _norm(value) -> str:
    return re.sub(r"\s+", " ", _clean(value).lower()).strip()


def _match_key(value) -> str:
    text = unicodedata.normalize("NFKD", _clean(value))
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"[^a-z0-9]", "", text.lower())


def _column(df: pd.DataFrame, aliases: list[str], required: bool = False) -> str | None:
    columns = {_norm(col): col for col in df.columns}
    for alias in aliases:
        key = _norm(alias)
        if key in columns:
            return columns[key]
    if required:
        raise ValueError(f"Required column missing for portal export: {aliases[0]}")
    return None


def _city_value(row: pd.Series) -> str:
    for column in row.index:
        header = _norm(column)
        if header == "city" or header.startswith("city.") or "delivery city" in header or header in {"wilayat", "wilayat name"}:
            value = _clean(row[column])
            if value:
                return value
    return ""


def _load_location_map(country: str) -> dict[str, str]:
    config = SETTINGS[country]
    path = Path(__file__).resolve().parent / "assets" / config["location_file"]
    mapping: dict[str, str] = {}
    if not path.exists():
        return mapping
    with path.open(newline="", encoding="utf-8-sig") as file:
        for row in csv.DictReader(file):
            area = _clean(row.get("AREA"))
            zone = _clean(row.get(config["zone_field"]))
            key = _match_key(area)
            if key and zone and key not in mapping:
                mapping[key] = zone
    return mapping


LOCATION_MAPS = {country: _load_location_map(country) for country in SETTINGS}


def _location_zone(country: str, city: str) -> int | str | None:
    key = _match_key(city)
    if not key:
        return None
    mapping = LOCATION_MAPS[country]
    zone = mapping.get(key)
    if not zone:
        candidates = [
            (area_key, value)
            for area_key, value in mapping.items()
            if area_key and (area_key in key or key in area_key)
        ]
        if candidates:
            candidates.sort(key=lambda item: len(item[0]), reverse=True)
            zone = candidates[0][1]
    if not zone:
        return None
    return int(zone) if zone.isdigit() else zone


def _digits(value) -> str:
    text = _clean(value)
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    return re.sub(r"\D", "", text)


def _phone(value, country: str) -> str:
    number = _digits(value)
    code = SETTINGS[country]["code"]
    if number.startswith("00" + code):
        number = number[len(code) + 2:]
    elif number.startswith(code):
        number = number[len(code):]
    elif number.startswith("0"):
        number = number[1:]
    return number if len(number) == 8 else ""


def _number(value) -> float:
    text = _clean(value).replace(",", "")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return float(match.group()) if match else 0.0


def _quantity(value):
    number = _number(value)
    if number <= 0:
        return 1
    return int(number) if number.is_integer() else number


def transform_gcc(source_df: pd.DataFrame, country: str) -> pd.DataFrame:
    if country not in SETTINGS:
        raise ValueError(f"Unsupported GCC country: {country}")

    cols = {
        key: _column(source_df, aliases, required=key in {"name", "country", "product"})
        for key, aliases in ALIASES.items()
    }

    mask = source_df[cols["country"]].map(
        lambda value: _norm(value) in SETTINGS[country]["countries"]
    )
    filtered = source_df[mask].copy()
    config = SETTINGS[country]
    records = []

    def get(row, key):
        column = cols.get(key)
        return row[column] if column else ""

    for index, row in filtered.iterrows():
        phone = _phone(get(row, "phone1"), country) or _phone(get(row, "phone2"), country)
        wilayat = _city_value(row)
        zone = _location_zone(country, wilayat)

        product1 = _clean(get(row, "product"))
        product2 = _clean(get(row, "product2"))
        qty1 = _quantity(get(row, "qty")) if product1 else 0
        qty2 = _quantity(get(row, "qty2")) if product2 else 0

        total_amount = _number(get(row, "amount"))
        total_qty = qty1 + qty2
        unit_price = total_amount / total_qty if total_qty else total_amount

        reference = _clean(get(row, "reference")) or f"{config['prefix']}{index + 2}"
        payment = _clean(get(row, "payment"))

        common = {
            "client_order_ref": reference,
            "customer_name": _clean(get(row, "name")),
            "partner_id": int(phone) if phone else None,
            "whatsapp_no": int(phone) if phone else None,
            "source_id": config["source_id"],
            "Pricelist Name": config["pricelist"],
            "street_no": _clean(get(row, "street")),
            "building_no": _clean(get(row, "building")) or None,
            "zone_id (governarate)": zone,
            "wilayat_id": wilayat or None,
            "city_id": None,
            "order_line/product_uom": "Units",
            "remarks": _clean(get(row, "remarks")) or None,
            "Agent Name": _clean(get(row, "agent_name")) or None,
            "Source": _clean(get(row, "source")) or None,
            "Payment Method": payment or None,
        }

        same_product = bool(product1 and product2 and _match_key(product1) == _match_key(product2))

        if same_product:
            record = common.copy()
            record.update({
                "order_line/product_id": product1,
                "order_line/price_unit": unit_price,
                "order_line/product_uom_qty": qty1 + qty2,
            })
            records.append(record)
            continue

        if product1:
            record1 = common.copy()
            record1.update({
                "order_line/product_id": product1,
                "order_line/price_unit": unit_price,
                "order_line/product_uom_qty": qty1,
            })
            records.append(record1)

        if product2:
            record2 = {header: None for header in HEADERS}
            record2.update({
                "order_line/product_id": product2,
                "order_line/product_uom": "Units",
                "order_line/price_unit": unit_price,
                "order_line/product_uom_qty": qty2,
                "Payment Method": payment or None,
            })
            records.append(record2)

    return pd.DataFrame(records, columns=HEADERS)


def _workbook_bytes(export_df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    duplicate_fill = PatternFill(fill_type="solid", fgColor="FFC7CE")

    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(1, col, header)
        cell.font = Font(name="Calibri", size=11, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    phones = export_df["partner_id"].fillna("").astype(str) if not export_df.empty else pd.Series(dtype=str)
    duplicate_mask = phones.ne("") & phones.duplicated(keep=False)

    for row_no, values in enumerate(export_df.itertuples(index=False, name=None), start=2):
        is_duplicate = bool(duplicate_mask.iloc[row_no - 2])
        for col_no, value in enumerate(values, start=1):
            cell = ws.cell(row_no, col_no, value)
            cell.font = Font(name="Calibri", size=11)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border
            if is_duplicate:
                cell.fill = duplicate_fill

    widths = [18, 24, 15, 15, 28, 24, 62, 18, 24, 24, 16, 34, 25, 24, 28, 34, 22, 22, 22]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    for row_no in range(2, ws.max_row + 1):
        ws.cell(row_no, 1).number_format = "@"
        ws.cell(row_no, 3).number_format = "0"
        ws.cell(row_no, 4).number_format = "0"
        ws.cell(row_no, 9).number_format = "0"
        ws.cell(row_no, 14).number_format = "0.00"
        ws.cell(row_no, 15).number_format = "0.##"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:S{max(1, ws.max_row)}"
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def build_gcc_export(source_df: pd.DataFrame, country: str) -> GCCResult:
    export_df = transform_gcc(source_df, country)
    return GCCResult(export_df, _workbook_bytes(export_df), len(export_df))
