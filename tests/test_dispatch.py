from pathlib import Path
import io
import pandas as pd
from openpyxl import load_workbook
from dispatch_core import build_export, read_workpex, transform

ROOT = Path(__file__).resolve().parents[1]
SOURCE = Path('/mnt/data/workpexexports.xlsx')
TEMPLATE = ROOT / 'assets' / 'UploadFormat.xlsx'


def test_filters_and_output():
    df = read_workpex(SOURCE)
    uae = transform(df, 'uae')
    oas = transform(df, 'oud_al_salam')
    assert len(uae) > 0
    assert 0 < len(oas) <= len(uae)
    with open(SOURCE, 'rb') as source, open(TEMPLATE, 'rb') as template:
        result = build_export(source, template, 'oud_al_salam')
    wb = load_workbook(io.BytesIO(result.workbook_bytes))
    ws = wb['Upload Format']
    assert ws.max_row == len(oas) + 1
    assert ws['A1'].value == 'CUSTOMER_NAME'
    assert ws['M1'].value == 'COLLECT_RETURN (YES/NO)'


def test_phone_fallback_and_payment_rules():
    source = pd.DataFrame([
        {
            'Lead Name': 'Primary UAE', 'Primary Phone': '971501234567',
            'Secondary Phone': '971509999999', 'Country': 'UAE',
            'Product': 'Test', 'Actual Amount': 120, 'Payment Method': 'COD'
        },
        {
            'Lead Name': 'Fallback UAE', 'Primary Phone': '919876543210',
            'Secondary Phone': 971506666666.0, 'Country': 'UAE',
            'Product': 'Test', 'Actual Amount': 130, 'Payment Method': 'Quick Link'
        },
        {
            'Lead Name': 'No UAE Phone', 'Primary Phone': '919876543210',
            'Secondary Phone': '966501234567', 'Country': 'UAE',
            'Product': 'Test', 'Actual Amount': 140, 'Payment Method': 'Account Transfer'
        },
    ])
    out = transform(source, 'uae')
    assert out.loc[0, 'MOBILE_NO'] == '501234567'
    assert out.loc[1, 'MOBILE_NO'] == '506666666'
    assert out.loc[2, 'MOBILE_NO'] == ''
    assert out.loc[0, 'COD_AMOUNT'] == 120
    assert out.loc[1, 'COD_AMOUNT'] == 0
    assert out.loc[2, 'COD_AMOUNT'] == 0
