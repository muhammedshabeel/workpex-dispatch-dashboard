from __future__ import annotations

import base64
import csv
import gzip
import io
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

BASE_DIR = Path(__file__).resolve().parent
ASSET_DIR = BASE_DIR / "assets"

NAQEL_PRODUCTS = {
    "ABSOLUTE MOUNTAIN AVENUE": "ABSOLUTE MOUNTAIN AVENUE_OUD AL SALAM",
    "ARBE PURO COMBO": "ARBE PURO COMBO_LPG",
    "LEON": "LEON_LPG",
    "OUD LOVERS": "OUD LOVERS_LPG",
    "PREMIUM EDITION": "PREMIUM COLLECTION_OUD AL SALAM",
    "PREMIUM COLLECTION": "PREMIUM COLLECTION_OUD AL SALAM",
}

ALIASES = {
    "name": ["Lead Name", "Customer Name", "Name", "First Name"],
    "phone1": ["Primary Phone", "Phone 1", "Phone1", "Phone", "Mobile", "Mobile No"],
    "phone2": ["Secondary Phone", "Phone 2", "Phone2", "Alternate Phone", "WhatsApp Number"],
    "country": ["Country"],
    "state": ["State", "Province", "Region"],
    "street": ["Street", "Address", "Address 1"],
    "city": ["CITY", "City", "Delivery City"],
    "national_code": ["National Code", "Reference No", "Order ID"],
    "product1": ["Product", "Product Name"],
    "qty1": ["QTY", "Quantity"],
    "product2": ["PRODUCT 2", "Product 2"],
    "qty2": ["QTY OF PRODUCT 2", "Quantity of Product 2", "QTY 2", "Qty 2"],
    "amount": ["Actual Amount", "Forecasted Amount", "Amount", "COD Amount"],
    "payment": ["Payment Method", "Payment"],
    "remarks": ["Lead Description", "Remarks", "Notes"],
    "agent": ["Assigned", "Assigned User"],
    "source": ["Source"],
}


@dataclass(frozen=True)
class KSAResult:
    naqel_df: pd.DataFrame
    jnt_df: pd.DataFrame
    naqel_bytes: bytes
    jnt_bytes: bytes
    naqel_orders: int
    jnt_orders: int


def _clean(value) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def _norm(value) -> str:
    return re.sub(r"\s+", " ", _clean(value).lower()).strip()


def _key(value) -> str:
    text = unicodedata.normalize("NFKD", _clean(value))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"[^a-z0-9]", "", text.lower())


def _column(df: pd.DataFrame, alias_key: str, required: bool = False) -> str | None:
    columns = {_norm(c): c for c in df.columns}
    for alias in ALIASES[alias_key]:
        if _norm(alias) in columns:
            return columns[_norm(alias)]
    if required:
        raise ValueError(f"Required Workpex column missing: {ALIASES[alias_key][0]}")
    return None


def _city_from_row(row: pd.Series) -> str:
    for column in row.index:
        header = _norm(column)
        if header.startswith("city") or "delivery city" in header:
            value = _clean(row[column])
            if value:
                return value
    return ""


def _digits(value) -> str:
    text = _clean(value)
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    return re.sub(r"\D", "", text)


def _ksa_phone(value) -> str:
    number = _digits(value)
    if number.startswith("00966"):
        number = number[5:]
    elif number.startswith("966"):
        number = number[3:]
    if number.startswith("0"):
        number = number[1:]
    return number if len(number) == 9 and number.startswith("5") else ""


def _number(value) -> float:
    text = _clean(value).replace(",", "")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return float(match.group()) if match else 0.0


def _qty(value, product_present: bool) -> float:
    if not product_present:
        return 0
    number = _number(value)
    return number if number > 0 else 1


def _template_bytes(name: str) -> bytes:
    xlsx = ASSET_DIR / name
    if xlsx.exists():
        return xlsx.read_bytes()
    gz_b64 = ASSET_DIR / f"{name}.gz.b64"
    if gz_b64.exists():
        return gzip.decompress(base64.b64decode(gz_b64.read_text(encoding="ascii")))
    b64 = ASSET_DIR / f"{name}.b64"
    if b64.exists():
        return base64.b64decode(b64.read_text(encoding="ascii"))
    raise FileNotFoundError(f"Missing KSA template asset: {name}")


def _load_prices() -> dict[str, float]:
    path = ASSET_DIR / "product_vendor.csv"
    prices: dict[str, float] = {}
    if not path.exists():
        return prices
    with path.open(encoding="utf-8-sig", newline="") as file:
        rows = list(csv.reader(file))
    for row in rows[2:]:
        if len(row) < 5:
            continue
        name = _clean(row[1])
        ksa = _clean(row[4])
        if not name or not ksa or "NOT AVAILABLE" in ksa.upper() or "TRANSIT" in ksa.upper():
            continue
        price = _number(ksa)
        if price > 0:
            prices[_key(name)] = price
    return prices


def _load_coverage() -> tuple[set[str], set[str]]:
    path = ASSET_DIR / "ksa_coverage.csv"
    naqel, jnt = set(), set()
    if not path.exists():
        return naqel, jnt
    with path.open(encoding="utf-8-sig", newline="") as file:
        rows = list(csv.reader(file))
    for row in rows[4:]:
        if len(row) > 3 and _clean(row[3]):
            naqel.add(_key(row[3]))
        if len(row) > 2 and _clean(row[2]):
            jnt.add(_key(row[2]))
    return naqel, jnt


PRICE_MAP = _load_prices()
NAQEL_COVERAGE, JNT_COVERAGE = _load_coverage()


def _canonical_product(raw: str) -> tuple[str, float, bool]:
    key = _key(raw)
    stripped = re.sub(r"(oudalsalam|lpg|atyaf|scentpassion|scntpassion)$", "", key)
    for base, portal in NAQEL_PRODUCTS.items():
        base_key = _key(base)
        if base_key in key or base_key == stripped:
            return portal, PRICE_MAP.get(base_key, 0.0), True

    canonical = _clean(raw).upper()
    jnt_aliases = {
        "HECTOR": "HECTOR COMBO",
        "ARCHER COMBO": "THE ARCHER COMBO",
        "VOLGA COMBO": "VOLGA EDITION PERFUME COMBO",
        "COLLECTION OF MOOD": "COLLECTION OF MOOD",
        "ASEEL COMBO": "ASEEL COMBO",
        "MIRAMAR": "MIRAMAR",
        "SHADOW FLAME": "SHADOW FLAME",
    }
    for base, portal in jnt_aliases.items():
        if _key(base) in key:
            canonical = portal
            break
    price = 0.0
    for price_key, value in PRICE_MAP.items():
        if price_key and price_key in key:
            price = value
            break
    return canonical, price, False


def _same_product(a: str, b: str) -> bool:
    return bool(a and b and _key(a) == _key(b))


def _city_maps():
    naqel_wb = load_workbook(io.BytesIO(_template_bytes("ksa_naqel_template.xlsx")), read_only=True, data_only=True)
    naqel_map: dict[str, tuple[str, str]] = {}
    if "Cities" in naqel_wb.sheetnames:
        ws = naqel_wb["Cities"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            code, _ar, en, _type = (list(row) + [None] * 4)[:4]
            if en:
                naqel_map[_key(en)] = (_clean(code), _clean(en))

    jnt_wb = load_workbook(io.BytesIO(_template_bytes("ksa_jnt_template.xlsx")), read_only=True, data_only=True)
    jnt_map: dict[str, tuple[str, str]] = {}
    if "Sheet1" in jnt_wb.sheetnames:
        ws = jnt_wb["Sheet1"]
        for row in ws.iter_rows(min_row=1, values_only=True):
            values = list(row) + [None] * 7
            city = _clean(values[1])
            province = _clean(values[5])
            for alias in [city, _clean(values[2]), _clean(values[3]), _clean(values[4])]:
                if alias:
                    jnt_map[_key(alias)] = (province, city)
    return naqel_map, jnt_map


NAQEL_CITY_MAP, JNT_CITY_MAP = _city_maps()


def _find_city(mapping: dict[str, tuple[str, str]], city: str, address: str = "") -> tuple[str, str]:
    city_key = _key(city)
    if city_key in mapping:
        return mapping[city_key]
    address_key = _key(address)
    matches = [(key, value) for key, value in mapping.items() if key and key in address_key]
    if matches:
        matches.sort(key=lambda item: len(item[0]), reverse=True)
        return matches[0][1]
    return "", city


def _route_order(product1: str, product2: str, city: str) -> str:
    _, _, naqel1 = _canonical_product(product1)
    _, _, naqel2 = _canonical_product(product2) if product2 else ("", 0.0, True)
    city_key = _key(city)
    city_covered = city_key in NAQEL_COVERAGE or city_key in NAQEL_CITY_MAP
    return "naqel" if naqel1 and naqel2 and city_covered else "jnt"


def _clear_data(ws, start_row: int = 2):
    if ws.max_row >= start_row:
        ws.delete_rows(start_row, ws.max_row - start_row + 1)


def _write_rows(ws, rows: list[list], start_row: int = 2):
    duplicate_fill = PatternFill(fill_type="solid", fgColor="FFC7CE")
    phones = [str(row[2] or "") for row in rows]
    duplicates = {phone for phone in phones if phone and phones.count(phone) > 1}
    for row_index, values in enumerate(rows, start=start_row):
        for column_index, value in enumerate(values, start=1):
            cell = ws.cell(row_index, column_index, value)
            if str(values[2] or "") in duplicates:
                cell.fill = duplicate_fill


def build_ksa_exports(source_df: pd.DataFrame) -> KSAResult:
    cols = {
        key: _column(source_df, key, required=key in {"name", "country", "product1"})
        for key in ALIASES
    }

    def get(row, key):
        column = cols.get(key)
        return row[column] if column else ""

    filtered = source_df[source_df[cols["country"]].map(
        lambda value: _norm(value) in {"saudi arabia", "ksa", "saudi"}
    )].copy()

    naqel_records: list[dict] = []
    jnt_records: list[dict] = []
    naqel_orders = 0
    jnt_orders = 0

    for index, row in filtered.iterrows():
        product1_raw = _clean(get(row, "product1"))
        product2_raw = _clean(get(row, "product2"))
        if not product1_raw:
            continue

        product1, price1, _ = _canonical_product(product1_raw)
        product2, price2, _ = _canonical_product(product2_raw) if product2_raw else ("", 0.0, False)
        qty1 = _qty(get(row, "qty1"), bool(product1_raw))
        qty2 = _qty(get(row, "qty2"), bool(product2_raw))
        total_amount = _number(get(row, "amount"))
        city = _city_from_row(row) or _clean(get(row, "state"))
        street = _clean(get(row, "street"))
        route = _route_order(product1_raw, product2_raw, city)
        phone = _ksa_phone(get(row, "phone1")) or _ksa_phone(get(row, "phone2"))
        backup = _ksa_phone(get(row, "phone2"))
        reference = _clean(get(row, "national_code")) or f"EMK{index + 2}"
        payment = _clean(get(row, "payment"))
        remarks = _clean(get(row, "remarks"))
        agent = _clean(get(row, "agent"))
        source = _clean(get(row, "source"))

        total_qty = qty1 + qty2
        fallback_unit = total_amount / total_qty if total_qty else total_amount
        unit1 = price1 or fallback_unit
        unit2 = price2 or fallback_unit

        lines = []
        if _same_product(product1, product2):
            lines.append((product1, qty1 + qty2, unit1))
        else:
            lines.append((product1, qty1, unit1))
            if product2:
                lines.append((product2, qty2, unit2))

        common = {
            "ref": reference,
            "name": _clean(get(row, "name")),
            "phone": phone,
            "backup": backup,
            "city": city,
            "street": street,
            "payment": payment,
            "remarks": remarks,
            "agent": agent,
            "source": source,
        }

        if route == "naqel":
            naqel_orders += 1
            destination, matched_city = _find_city(NAQEL_CITY_MAP, city, street)
            for line_no, (product, quantity, unit_price) in enumerate(lines):
                naqel_records.append({
                    **common,
                    "dest": destination,
                    "matched_city": matched_city,
                    "product": product,
                    "qty": quantity,
                    "unit_price": unit_price,
                    "line_no": line_no,
                })
        else:
            jnt_orders += 1
            province, matched_city = _find_city(JNT_CITY_MAP, city, street)
            for line_no, (product, quantity, unit_price) in enumerate(lines):
                jnt_records.append({
                    **common,
                    "province": province,
                    "matched_city": matched_city,
                    "product": product,
                    "qty": quantity,
                    "unit_price": unit_price,
                    "line_no": line_no,
                })

    naqel_wb = load_workbook(io.BytesIO(_template_bytes("ksa_naqel_template.xlsx")))
    naqel_ws = naqel_wb["GenerateWaybills"]
    _clear_data(naqel_ws, 2)
    for row_index, record in enumerate(naqel_records, start=2):
        declared = record["unit_price"] * record["qty"]
        values = [
            record["ref"], "RUH", record["dest"], record["name"], None,
            record["phone"], record["phone"], record["matched_city"], record["street"],
            None, None, None, None, record["qty"], 1.5, 10, 10, 10, declared,
            record["remarks"] or "NIL", "NIL", declared, "SAR", "OTHER/UNKNOWN",
            record["product"], None, None, None, None, None, "KSA", "KSA",
        ]
        for column_index, value in enumerate(values, start=1):
            naqel_ws.cell(row_index, column_index, value)
    naqel_output = io.BytesIO()
    naqel_wb.save(naqel_output)

    jnt_wb = load_workbook(io.BytesIO(_template_bytes("ksa_jnt_template.xlsx")))
    jnt_ws = jnt_wb["Template"]
    _clear_data(jnt_ws, 2)
    extra_headers = ["Agent Name", "Source", "Payment Method", "Quantity", "Unit Price"]
    for index, header in enumerate(extra_headers, start=29):
        jnt_ws.cell(1, index, header)

    jnt_rows = []
    for record in jnt_records:
        amount = record["unit_price"] * record["qty"]
        jnt_rows.append([
            record["ref"], record["name"], int(record["phone"]) if record["phone"] else None,
            int(record["backup"]) if record["backup"] else None, record["province"],
            record["matched_city"], None, None, record["street"], None, None, None,
            "HOSA5275", None, None, "STANDARD",
            "Cash" if _norm(record["payment"]) in {"cod", "cash on delivery"} else record["payment"],
            None, "Others", 1, record["product"], "No", None, None, None, amount,
            "NO", record["remarks"], record["agent"], record["source"], record["payment"],
            record["qty"], record["unit_price"],
        ])
    _write_rows(jnt_ws, jnt_rows, 2)
    jnt_output = io.BytesIO()
    jnt_wb.save(jnt_output)

    return KSAResult(
        pd.DataFrame(naqel_records),
        pd.DataFrame(jnt_records),
        naqel_output.getvalue(),
        jnt_output.getvalue(),
        naqel_orders,
        jnt_orders,
    )
